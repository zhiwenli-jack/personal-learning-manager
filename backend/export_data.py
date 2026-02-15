"""数据导出脚本 - 将数据库中的数据导出为 Excel"""
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, ForeignKey, Boolean, Enum, JSON, Numeric
from sqlalchemy.orm import declarative_base, sessionmaker
from datetime import datetime
import json


# 数据库路径
DB_PATH = Path(__file__).parent / "personal_study.db"
DATABASE_URL = f"sqlite:///{DB_PATH}"

# 创建基础类
Base = declarative_base()


class Direction(Base):
    """学习方向表"""
    __tablename__ = "directions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(50))
    description = Column(Text)
    created_at = Column(DateTime)


class Material(Base):
    """学习资料表"""
    __tablename__ = "materials"
    id = Column(Integer, primary_key=True, autoincrement=True)
    direction_id = Column(Integer, ForeignKey("directions.id"))
    title = Column(String(200))
    content = Column(Text)
    key_points = Column(JSON)
    status = Column(String(20))
    created_at = Column(DateTime)

    direction = None


class Question(Base):
    """题目表"""
    __tablename__ = "questions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    material_id = Column(Integer, ForeignKey("materials.id"))
    type = Column(String(20))
    difficulty = Column(Integer)
    content = Column(Text)
    options = Column(JSON)
    answer = Column(Text)
    explanation = Column(Text)
    rating = Column(String(10))
    created_at = Column(DateTime)

    material = None


class Answer(Base):
    """答题记录表"""
    __tablename__ = "answers"
    id = Column(Integer, primary_key=True, autoincrement=True)
    exam_id = Column(Integer)
    question_id = Column(Integer, ForeignKey("questions.id"))
    user_answer = Column(Text)
    is_correct = Column(Boolean)
    score = Column(Numeric(5, 2))
    ai_feedback = Column(Text)
    answered_at = Column(DateTime)


class Mistake(Base):
    """错题表"""
    __tablename__ = "mistakes"
    id = Column(Integer, primary_key=True, autoincrement=True)
    question_id = Column(Integer, ForeignKey("questions.id"))
    answer_id = Column(Integer, ForeignKey("answers.id"))
    review_count = Column(Integer)
    mastered = Column(Boolean)
    created_at = Column(DateTime)

    question = None
    answer = None


class ParseTask(Base):
    """知识解析任务表"""
    __tablename__ = "parse_tasks"
    id = Column(Integer, primary_key=True, autoincrement=True)
    direction_id = Column(Integer, ForeignKey("directions.id"))
    title = Column(String(200))
    source_type = Column(String(10))
    source_content = Column(Text)
    raw_text = Column(Text)
    summary = Column(Text)
    status = Column(String(20))
    error_message = Column(Text)
    created_at = Column(DateTime)
    updated_at = Column(DateTime)

    direction = None


def get_db_session():
    """获取数据库会话"""
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
    Session = sessionmaker(bind=engine)
    return Session()


def style_header(ws, max_col: int):
    """设置表头样式"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def export_parse_tasks(session, wb: Workbook):
    """导出解析记录"""
    ws = wb.create_sheet("解析记录")

    # 表头
    headers = ["ID", "任务标题", "学习方向", "来源类型", "来源内容", "摘要", "状态", "创建时间", "更新时间"]
    ws.append(headers)
    style_header(ws, len(headers))

    # 查询数据
    tasks = session.query(ParseTask).order_by(ParseTask.created_at.desc()).all()

    for task in tasks:
        direction_name = task.direction.name if task.direction else ""
        ws.append([
            task.id,
            task.title,
            direction_name,
            task.source_type if task.source_type else "",
            task.source_content or "",
            task.summary or "",
            task.status if task.status else "",
            task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else "",
            task.updated_at.strftime("%Y-%m-%d %H:%M:%S") if task.updated_at else "",
        ])

    # 设置列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18

    print(f"  解析记录: {len(tasks)} 条")


def export_questions(session, wb: Workbook):
    """导出题目管理"""
    ws = wb.create_sheet("题目管理")

    # 表头
    headers = ["ID", "题目内容", "题目类型", "难度", "选项", "标准答案", "答案解析", "用户评价", "创建时间", "资料来源"]
    ws.append(headers)
    style_header(ws, len(headers))

    # 查询数据
    questions = session.query(Question).order_by(Question.created_at.desc()).all()

    for q in questions:
        # 处理选项格式
        options_str = ""
        if q.options:
            if isinstance(q.options, dict):
                options_str = "\n".join([f"{k}. {v}" for k, v in q.options.items()])
            elif isinstance(q.options, list):
                options_str = "\n".join(q.options)

        material_title = q.material.title if q.material else ""

        ws.append([
            q.id,
            q.content,
            q.type if q.type else "",
            q.difficulty,
            options_str,
            q.answer,
            q.explanation or "",
            q.rating if q.rating else "",
            q.created_at.strftime("%Y-%m-%d %H:%M:%S") if q.created_at else "",
            material_title,
        ])

    # 设置列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 25

    # 设置行高
    for row in range(2, len(questions) + 2):
        ws.row_dimensions[row].height = 40

    print(f"  题目管理: {len(questions)} 条")


def export_mistakes(session, wb: Workbook):
    """导出错题本"""
    ws = wb.create_sheet("错题本")

    # 表头
    headers = ["ID", "错题内容", "题目类型", "标准答案", "用户答案", "是否正确", "AI反馈", "复习次数", "是否已掌握", "错题时间", "资料来源"]
    ws.append(headers)
    style_header(ws, len(headers))

    # 查询数据 - 直接用 SQL 查询获取关联信息
    query = """
        SELECT m.id, q.content, q.type, q.answer, a.user_answer, a.is_correct,
               a.ai_feedback, m.review_count, m.mastered, m.created_at, mat.title as material_title
        FROM mistakes m
        LEFT JOIN questions q ON m.question_id = q.id
        LEFT JOIN answers a ON m.answer_id = a.id
        LEFT JOIN materials mat ON q.material_id = mat.id
        ORDER BY m.created_at DESC
    """
    result = session.execute(query)

    for row in result:
        # 处理时间字段
        created_at = row[9]
        if isinstance(created_at, str):
            created_at_str = created_at
        elif created_at:
            created_at_str = created_at.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_at_str = ""

        ws.append([
            row[0],
            row[1] or "",
            row[2] or "",
            row[3] or "",
            row[4] or "",
            "是" if row[5] else "否",
            row[6] or "",
            row[7],
            "是" if row[8] else "否",
            created_at_str,
            row[10] or "",
        ])

    # 设置列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 25

    # 设置行高
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 40

    print(f"  错题本: {ws.max_row - 1} 条")


def main():
    """主函数"""
    print("=" * 50)
    print("数据导出工具")
    print("=" * 50)

    session = get_db_session()

    try:
        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认sheet

        # 导出各模块数据
        print("\n正在导出数据...")
        export_parse_tasks(session, wb)
        export_questions(session, wb)
        export_mistakes(session, wb)

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"学习数据导出_{timestamp}.xlsx"
        output_path = Path(__file__).parent / filename

        # 保存文件
        wb.save(output_path)

        print(f"\n✅ 导出成功！")
        print(f"   文件路径: {output_path}")
        print(f"   文件大小: {output_path.stat().st_size / 1024:.1f} KB")

    finally:
        session.close()


if __name__ == "__main__":
    main()
